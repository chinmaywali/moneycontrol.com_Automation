
import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.max_column


def readData(file, sheetName, rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    return sheet.cell(rowNum, colNum).value


def writeData(file, sheetName, rowNum, colNum, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    sheet.cell(rowNum, colNum).value = data
    wb.save(file)


def fillGreenColor(file, sheetName, rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    sheet = wb[sheetName]
    greenFill = PatternFill(start_color="60b212", end_color="60b212", fill_type='solid')
    sheet.cell(rowNum,colNum).fill = greenFill
    wb.save(file)
